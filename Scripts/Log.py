"""
Whenever a model is trained it should have the weights saved and some descriptive information readily available
in a document, so that it is possible to know the settings used for the model.

Data used for training should be designated, and feature augmented data should be stored in /Augmented_Data

"""
from pathlib import Path
import json
from PIL import Image
import numpy as np
from git import Repo
from matplotlib.pyplot import hist
import numpy as np
from numpy.linalg import norm
from matplotlib.colors import Normalize, BoundaryNorm
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from pandas import read_csv, DataFrame
from keras.models import load_model
from pandas import read_excel
import segyio
from NGI.GM_Toolbox import evaluate_modeldist

msc_color = '#6bb4edff'
dark_msc_color = '#1466a6ff'
coral = '#f87954ff'
dark_coral = '#c93408ff'

def gname(old_name):
    """ Takes in the state of group names and outputs the next one
    """
    if not len(old_name): raise TypeError('Max group name encountered')
    new_name = old_name
    if old_name[-1] != 'Z':
        new_name = old_name[:-1] + chr(ord(old_name[-1])+1)
    else:
        new_name = gname(old_name[:-1]) + 'A'
    return new_name

def update_groupstate(gname):
    """Keeps groupstate updated after any model run
    """
    fp = 'Models/_groupstate.json'
    message = 'Modelrun {} automatic push'.format(gname)
    repo_push(fp, message)

def new_group(push=True):
    """
    Creates a log of the model being run and puts it in the \Models 
    directory. Uses the RunModel object to extract the needed values
    """    
    with open('./Models/_groupstate.json', 'r+') as state:
        group = json.loads(state.read())
        state.seek(0)
        group['Group'] = gname(group['Group'])
        group_name = group['Group']

        state.write(json.dumps(group))
    

    # Creates parent directory
    m = Path('./Models')
    new = m.joinpath(group_name)
    Path.mkdir(new, parents=True, exist_ok=True)
    
    if push: update_groupstate(group['Group'])
    return group_name

def nats(k):
    yield k
    yield from nats(k+1)

def give_modelname():
    n = nats(0)
    groupname = new_group()
    while True: # Always
        yield groupname, next(n)

def repo_push(fps, message):
    """
    Automatically pushes selected files to repo
    with an automatic commit message.
    """
    try:
        repo = Repo('.')
        for fp in [fps]:
            repo.git.add(fp)
        repo.index.commit(message)
        origin = repo.remote(name = 'origin')
        origin.push()
    except:
        print('Unable to push to remote repo')

def plot_history(History, filename=None):
    """Plots the history of the model training"""
    for key in History.history.keys():
        # if 'loss' in key:
        #     if key == 'loss':
        #         plt.plot(History.history[key], label=key, color='k', linewidth=2, zorder=2)
        #     else:
        if ('decoder' in key) and ('loss' in key):    
            plt.plot(History.history[key], label=key, c=msc_color, linewidth=2, zorder=1)
    plt.yscale('log')
    plt.title('Model loss', fontsize=20)
    plt.ylabel('Loss')
    plt.xlabel('Epoch')
    plt.legend(loc='upper right')
    plt.savefig(filename)
    plt.close()

def plot_reconstruction(reconstruct, cpt_loc=15, zrange = (35, 60), filename = ''):
    d = read_excel('../OneDrive - NGI/Documents/NTNU/MSC_DATA/Distances_to_2Dlines_Revised.xlsx')
    d = d.loc[d['Location no.'] == cpt_loc]
    # Pick the first of the d
    d = d.iloc[0]
    # Get the 2D UHR line
    line = d['2D UHR line']
    CDP = d['CDP']

    seis_file = '../OneDrive - NGI/Documents/NTNU/MSC_DATA/2DUHRS_06_MIG_DEPTH/{}.sgy'.format(line)

    with segyio.open(seis_file, ignore_geometry=True) as f:
        samples = f.samples
        seis = segyio.collect(f.trace.raw[:])
        cdps = segyio.collect(f.attributes(segyio.TraceField.CDP)[:])
        # Get the index of the CDP
        cdp_idx = np.where(abs(cdps - CDP)<600)[0]
        seis = seis[cdp_idx][:, (samples >= zrange[0]) & (samples < zrange[1])]
    
    # Reshape the seismic into images of 11, depth, 1
    X_seis = seis.reshape((-1, 11, seis.shape[1], 1))

    pred = reconstruct.predict(X_seis)[1]
    pred = pred.reshape((-1, seis.shape[1]))

    fig, ax = plt.subplots(1, 2, figsize=(10, 5))
    
    ax[0].imshow(seis.T, aspect='auto', cmap='gray')
    ax[0].xaxis.set_visible(False)
    # Set y values to depth
    ax[0].set_yticks(np.arange(0, seis.shape[1], 150))
    ax[0].set_yticklabels(samples[(samples >= zrange[0]) & (samples < zrange[1])][::150])
    ax[0].set_ylabel('Depth [mLAT]', fontsize=15)

    ax[0].set_title('Original', fontsize=20)
    ax[1].imshow(pred.T, aspect='auto', cmap='gray')
    ax[1].xaxis.set_visible(False)
    ax[1].yaxis.set_visible(False)
    ax[1].set_title('Reconstruction', fontsize=20)
    fig.tight_layout()

    

    if filename:
        plt.savefig(filename)
    else:
        plt.show()
    plt.close()
    
    # Create a histogram of the predicted values and the true values

    fig, ax = plt.subplots(1, 1, figsize=(10, 5))
    ax.hist(pred.flatten(), bins=50, label='Predicted', alpha=0.5)
    ax.hist(seis.flatten(), bins=50, label='True', alpha=0.5)
    ax.legend()
    ax.set_title('Histogram of predicted values', fontsize=20)
    ax.set_xlabel('Predicted value', fontsize=15)
    ax.set_ylabel('Count', fontsize=15)
    fig.tight_layout()
    if filename:
        plt.savefig(filename.replace('.png', '_hist.png'))
    else:
        plt.show()
    plt.close()

def plot_dontfit(filename='', zrange = (35, 60)):
    from Feat_aug import create_full_trace_dataset, get_cpt_name

    Full_args = create_full_trace_dataset(zrange=zrange)

    y = Full_args[1]
    z = np.arange(zrange[0], zrange[1], 0.1)
    groups = Full_args[2]
    minmax = Full_args[-1]
    ggm = Full_args[7]

    cmap, norm, _ = get_GGM_cmap(ggm)
    umap = get_umap_func()

    locations = [3, 6, 7, 33, 71, 72]

    params = ['q_c', 'f_s', 'u_2']
    colors = ['g', 'orange', 'b']

    fig, ax = plt.subplots(4, 8, figsize=(20, 30))
    fig.subplots_adjust(left = 0.076, bottom=0.05, right=0.92, top=0.93, wspace=0.32, hspace=0.3)

    lens = np.array([len(y[np.where(groups==cpt_loc)]) for cpt_loc in locations])[:-1]-1
    i_list = np.arange(0, len(locations))
    i_list[1:] = i_list[1:] + np.cumsum(lens)

    for i, cpt_loc in zip(i_list, locations):
        cptname = get_cpt_name(cpt_loc)
        y_loc = y[np.where(groups == cpt_loc)]
        mins, maxs = (minmax[0][np.where(groups == cpt_loc)], minmax[1][np.where(groups == cpt_loc)])
        g = ggm[np.where(groups == cpt_loc)]

        for ii in range(i, y_loc.shape[0]+i):
            for j in range(3):
                if j == 0: ax[ii%4, j+4*(ii//4)].set_ylabel(cptname)
                if j == 2: ax[ii%4, j+4*(ii//4)].set_ylabel('Depth (m)'); ax[ii%4, j+4*(ii//4)].yaxis.set_label_position("right"); ax[ii%4, j+4*(ii//4)].yaxis.tick_right()
                else: ax[ii%4, j+4*(ii//4)].set_yticks([])
                ax[ii%4, j+4*(ii//4)].plot(y_loc[(ii-i)%4, :, j], z, color=colors[j], label=params[j])
                ax[ii%4, j+4*(ii//4)].fill_betweenx(z, mins[(ii-i)%4, :, j], maxs[(ii-i)%4, :, j], alpha=0.5, color=colors[j])
                ax[ii%4, j+4*(ii//4)].invert_yaxis()
                if ii%4 == 3: ax[ii%4, j+4*(ii//4)].set_xlabel(f'${params[j]}$')
            #    get the limits of the plots in row i

            # set the first three plots in the row to the same y axis limits
            ylims = [ax[ii%4, j+4*(ii//4)].get_ylim() for j in range(3)]
            ylims = np.max(ylims, axis=0)

            # set the limit of the plots in row i to the max of the limits in row i
            gm = g[int(ii-i)][np.where((z>=ylims[1])&(z<=ylims[0]))]
            gm_z = z[np.where((z>=ylims[1])&(z<=ylims[0]))]
            
            ax[ii%4, 3 + 4*(ii//4)].imshow(gm.reshape(-1, 1), cmap=cmap, norm=norm, extent=[gm_z[0], gm_z[-1], gm_z[-1], gm_z[0]], aspect=8)
            ylims = ax[ii%4, 3 + 4*(ii//4)].get_ylim()
            [ax[ii%4, j+4*(ii//4)].set_ylim(ylims) for j in range(3)]
            changes = np.diff(gm)
            depth_changes = gm_z[np.where(changes != 0)]
            depth_changes = np.insert(depth_changes, 0, gm_z[0]); depth_changes = np.append(depth_changes, ylims[0])
            depth_changes = np.sort(depth_changes)
            diff_changes = np.diff(depth_changes)
            y_ticks = depth_changes[:-1] + diff_changes/2
            
            yticklabels = [umap(x) for x in np.unique(gm.flatten())]
            ax[ii%4, 3 + 4*(ii//4)].yaxis.tick_right()
            ax[ii%4, 3 + 4*(ii//4)].set_yticks(y_ticks)
            ax[ii%4, 3 + 4*(ii//4)].set_yticklabels(yticklabels, rotation=20, ha='left', va='bottom', fontsize=6)
            ax[ii%4, 3 + 4*(ii//4)].set_xticks([])
        

    # plt.legend()
    fig.suptitle('Discarded CPT measurements', fontsize=16)
    if filename:
        plt.savefig(filename, dpi=500)
    else:
        plt.show()

def get_GGM_cmap(GGM):
    """Creates a segmented colorbar with units colored by their uid"""

    umap = read_csv('../OneDrive - NGI/Documents/NTNU/MSC_DATA/StructuralModel_unit_mapping.csv')
    GGM_names = []

    unique_uid = np.sort(np.unique(GGM.flatten())).astype(int)

    for u in unique_uid:
        GGM_names.append(umap.loc[umap['uid'] == u]['unit'].values[0])

    n_colors = len(np.unique(umap['uid']))
    # Add a segmented colorbar with unique colors for the different units
    cmap = plt.cm.get_cmap('gnuplot', n_colors)

    # Define the bins and normalize
    bounds = np.linspace(0, n_colors, n_colors+1)
    norm = BoundaryNorm(bounds, cmap.N)

    # Create a colorbar with only the unique GGM provided to the function given the right colors
    # sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    # sm.set_array([])
    # cbar = plt.colorbar(sm, ticks=np.arange(len(unique_uid))+0.5)
    # cbar.ax.set_yticklabels(GGM_names)
    cbar = None
    return cmap, norm, cbar

def describe_data(X, y, groups, GGM, mdir=''):
    """Creates a text file with the description of the data"""

    unit_mapping = read_csv('../OneDrive - NGI/Documents/NTNU/MSC_DATA/StructuralModel_unit_mapping.csv', index_col='uid')

    filename = mdir + 'dataset_description.csv'
    latex_filename = mdir + 'dataset_description.tex'

    tex_string = ''
    data_df = DataFrame(columns=['qc', 'fs', 'u2', 'GGM'])

    # Make histogram of the GGM counts with unit mapping on the x axis centered at the bars
    fig, ax = plt.subplots()

    umap = lambda x: unit_mapping['unit'][int(x)]
    
    # Make a flattened sorted array of the GGM values
    ggm_flatsort = np.sort(GGM.flatten())
    # ax.hist(string_GGM, bins=np.arange(len(np.unique(GGM))+1), color=msc_color, edgecolor='k')
    unique_units = np.unique(ggm_flatsort, return_counts=True)
    # Add a bar plot with the counts of the GGM values
    ax.bar(np.arange(len(unique_units[0])), unique_units[1], color=msc_color, edgecolor='k')
    ax.set_xticks(np.arange(len(unique_units[0])))
    ax.set_xticklabels([umap(x) for x in unique_units[0]])
    # Rotate the xtick labels
    plt.setp(ax.get_xticklabels(), rotation=90, rotation_mode="anchor", ha="right", va="center")
    ax.set_xlabel('Unit')
    ax.set_ylabel('Count')

    # set y axis to log scale
    ax.set_yscale('log')

    fig.subplots_adjust(bottom=0.2)
    fig.suptitle('CPT values per GGM unit')
    fig.savefig(mdir + 'GGM_histogram.png', dpi=500)
    plt.close(fig)

    
    ggm = GGM.reshape(*y.shape[:-1], 1)
    flat_data = np.concatenate((y, ggm), axis=-1).reshape(-1, 4)

    df = DataFrame(flat_data, columns=['$q_c$', '$f_s$', '$u_2$', 'GGM'])
    df.to_csv(mdir+'data.csv')

    for g in df.groupby('GGM'):
        unit = umap(g[0])

        g_data = g[1].iloc[:, :-1]
        
        desc = g_data.describe().T
        desc.columns = ['n', '$\mu$', '$\sigma$', 'min', 'Q1', 'Q2', 'Q3', 'max']

        beg = '\\begin{table}[h]\n'
        string = beg + f'\\caption{{Summary statistics for {unit}}}\n'
        string += desc.to_latex(escape=False)
        string += '\\end{table}\n\n'
        tex_string += string

        data_df = data_df.append(g[1].describe())

    data_df.to_csv(filename)
    with open(latex_filename, 'w') as f:
        f.write(tex_string)

def get_umap_func():

    unit_mapping = read_csv('../OneDrive - NGI/Documents/NTNU/MSC_DATA/StructuralModel_unit_mapping.csv', index_col='uid')

    def umap(uid):
        return unit_mapping['unit'][int(uid)]

    return umap

def make_cv_excel(filename, COMP_DF):
    """Creates a new excel file with the results of the model"""
    xl = "Results/NGI_stdd_{}.xlsx"
    wb = load_workbook(xl)
    
    params = ['q_c', 'f_s', 'u_2']
    
    unit_mapping = read_csv('../OneDrive - NGI/Documents/NTNU/MSC_DATA/StructuralModel_unit_mapping.csv', index_col='uid')
    
    for p in params:
        ws = wb[p]
        param = p.replace('_', '')
        for method in ['CNN', 'RF', 'LGBM']:
            true = COMP_DF['True_{}'.format(param)]
            pred = COMP_DF['{}_{}'.format(method, param)]
            std = evaluate_modeldist(true, pred)[4]

            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == 'all':
                    cellrow = row
                    break

            for col in range(2, ws.max_column + 1):
                if '_'+method in ws.cell(row=1, column=col).value:
                    cellcol = col
                    break

            ws.cell(row=cellrow, column=cellcol).value = std


    for g in COMP_DF.groupby('GGM'):
        unit = unit_mapping['unit'][int(g[0])]
        g_data = g[1]

        for p in params:
            ws = wb[p]
            param = p.replace('_', '')
            for method in ['CNN', 'RF', 'LGBM']:
                true = g_data['True_{}'.format(param)]
                pred = g_data['{}_{}'.format(method, param)]
                std = evaluate_modeldist(true, pred)[4]
                
                # Find the cell row by searching for the GGM in the first column
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == unit:
                        cellrow = row
                        break
                # Find the cell column by searching for the method in the first row
                for col in range(2, ws.max_column + 1):
                    if '_'+method in ws.cell(row=1, column=col).value:
                        cellcol = col
                        break
                
                # set the cell value to the std
                ws.cell(row=cellrow, column=cellcol).value = std

    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    wb.save(filename)

def make_cv_excel_bestworst(filename, COMP_DF):
    """ Creates an excel file with the prediction std each for the best and worst case"""

    COMP_DF = COMP_DF.dropna(subset=['true_qc', 'true_fs', 'true_u2', 'GGM', 'TRUE_qc', 'TRUE_fs', 'TRUE_u2'])

    umap = get_umap_func()
    xl = "Results/NGI_stdd_{}.xlsx"
    # load a copy of the excel file

    wb_best = load_workbook(xl)

    wb_worst = load_workbook(xl)

    params = ['q_c', 'f_s', 'u_2']

    for p in params:
        ws = wb_best[p]
        ws_worst = wb_worst[p]
        param = p.replace('_', '')
        for method in ['CNN', 'RF', 'LGBM']:
            TRUE = COMP_DF['TRUE_{}'.format(param)]
            true = COMP_DF['true_{}'.format(param)]
            pred = COMP_DF['{}_{}'.format(method, param)]

            # pred = pred[~np.isnan(TRUE) & ~np.isnan(true)]
            # TRUE = TRUE[~np.isnan(TRUE)]
            # true = true[~np.isnan(true)]
            

            assert TRUE.shape == true.shape == pred.shape

            # Pick out the values of TRUE and true which gives the best and worst std
            DIFF = np.absolute(pred - TRUE)
            diff = np.absolute(pred - true)
            
            d = np.stack((DIFF, diff))
            # d = d[:, ~np.isnan(d).any(axis=0)]
            min_d = np.amin(d, axis=0)
            max_d = np.amax(d, axis=0)
            best_std = np.nanstd(min_d)
            worst_std = np.nanstd(max_d)


            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == 'all':
                    cellrow = row
                    break

            for col in range(2, ws.max_column + 1):
                if '_'+method in ws.cell(row=1, column=col).value:
                    cellcol = col
                    break

            ws.cell(row=cellrow, column=cellcol).value = best_std
            ws_worst.cell(row=cellrow, column=cellcol).value = worst_std
    
    for g in COMP_DF.groupby('GGM'):
        g_data = g[1]

        for p in params:
            ws = wb_best[p]
            ws_worst = wb_worst[p]
            param = p.replace('_', '')
            for method in ['CNN', 'RF', 'LGBM']:
                TRUE = g_data['TRUE_{}'.format(param)]
                true = g_data['true_{}'.format(param)]
                pred = g_data['{}_{}'.format(method, param)]

                # pred = pred[~np.isnan(TRUE) & ~np.isnan(true)]
                # TRUE = TRUE[~np.isnan(TRUE)]
                # true = true[~np.isnan(true)]
            

                assert TRUE.shape == true.shape == pred.shape

                # Pick out the values of TRUE and true which gives the best and worst std
                DIFF = np.absolute(pred - TRUE) # Difference for upper and lower bound
                diff = np.absolute(pred - true)
                
                d = np.stack((DIFF, diff))
                # d = d[:, ~np.isnan(d).any(axis=0)]
                min_d = np.amin(d, axis=0)
                max_d = np.amax(d, axis=0)
                best_std = np.nanstd(min_d)
                worst_std = np.nanstd(max_d)

                # Find the cell row by searching for the GGM in the first column
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == umap(g[0]):
                        cellrow = row
                        break
                # Find the cell column by searching for the method in the first row
                for col in range(2, ws.max_column + 1):
                    if '_'+method in ws.cell(row=1, column=col).value:
                        cellcol = col
                        break
                
                # set the cell value to the std
                ws.cell(row=cellrow, column=cellcol).value = best_std
                ws_worst.cell(row=cellrow, column=cellcol).value = worst_std

    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    wb_best.save(filename.replace('.xlsx', '_best.xlsx'))
    wb_worst.save(filename.replace('.xlsx', '_worst.xlsx'))

def add_identity(axes, *line_args, **line_kwargs):
    """Author: JaminSore (https://stackoverflow.com/users/1940479/jaminsore)"""
    identity, = axes.plot([], [], *line_args, **line_kwargs)
    def callback(axes):
        low_x, high_x = axes.get_xlim()
        low_y, high_y = axes.get_ylim()
        low = max(low_x, low_y)
        high = min(high_x, high_y)
        identity.set_data([low, high], [low, high])
    callback(axes)
    axes.callbacks.connect('xlim_changed', callback)
    axes.callbacks.connect('ylim_changed', callback)
    return axes

def loss_dict_latex(loss_dict, destination_folder='', filename='loss_dict_new'):
    """Creates a latex table from a loss dictionary"""
    from Feat_aug import get_cpt_name
    folds = np.unique([f[:f.find('_')] for f in loss_dict.keys() if f.startswith('Fold')])

    # Get parent folder of loss dict
    gname = 'Models/' + Path(destination_folder).name
    
    

    df = DataFrame([], columns = ['Fold', 'LOO', 'MAE_decoder', 'MSE_decoder', 'MAE_Rec', 'MSE_Rec'])

    for fold in folds:
        tempdict = {}
        tempdict['Fold'] = fold[4:]

        with open(gname + '/' + fold + '/LOO_group.txt', 'r') as f:
            loo = int(f.read()[11:].strip('[]'))

        tempdict['LOO'] = get_cpt_name(loo)

        for key in loss_dict.keys():
            if key.startswith(fold):
                if 'decoder' in key:
                    if 'mae' in key:
                        tempdict['MAE_decoder'] = loss_dict[key]
                    elif 'mse' in key:
                        tempdict['MSE_decoder'] = loss_dict[key]
                elif 'Rec' in key:
                    if 'mae' in key:
                        tempdict['MAE_Rec'] = loss_dict[key]
                    elif 'mse' in key:
                        tempdict['MSE_Rec'] = loss_dict[key]
        
        df = df.append(tempdict, ignore_index=True)

    df.to_excel(destination_folder+filename+'.xlsx', index=False)
    string = df.to_latex(escape=False, index=False)
    return string

def bar_plot_ggm(GGM, groups, filename=''):
    """Creates a bar plot with all occurrences of each group in the GGMs"""
    from Feat_aug import get_cpt_name
    umap = get_umap_func()
    unique_ggm = np.sort(np.unique(GGM))
    unique_groups = np.sort(np.unique(groups))
    ggm_counts = np.zeros((len(unique_ggm), len(unique_groups)))

    tmp = np.ones_like(GGM)
    for i, g in enumerate(groups):
        tmp[i] = np.ones_like(GGM[i])*g
    groups = tmp

    GGM = GGM.flatten()
    groups = groups.flatten()

    columns = ['GGM']*len(unique_ggm)
    rows = ['Group']*len(unique_groups)
    for i, ggm in enumerate(unique_ggm):
        columns[i] = umap(ggm)
        for j, group in enumerate(unique_groups):
            ggm_counts[i,j] = np.sum((GGM == ggm) & (groups == group))
            rows[j] = get_cpt_name(group)


    fig, ax = plt.subplots(figsize=(10, 5))
    for i, group in enumerate(unique_groups):
        ax.bar(np.arange(len(unique_ggm)), ggm_counts[:,i], bottom=np.sum(ggm_counts[:,:i], axis=1), label=get_cpt_name(group), zorder=3)

    # Set bar label to be the number of containers in each GGM
    bar_labels = [len(l[l!=0]) for l in ggm_counts]
    ax.bar_label(ax.containers[-1], labels=bar_labels, zorder=4, padding=3)
    # increase room over the bars
    ax.set_ylim(top=ax.get_ylim()[1]*1.1)
    ax.set_xticks(np.arange(len(unique_ggm)))
    ax.set_xticklabels([umap(x) for x in unique_ggm], rotation=45, ha='right')
    ax.set_ylabel('Number of samples')
    # Set horizontal grid lines
    ax.yaxis.grid(True, zorder=1)

    fig.suptitle('Distribution of GGM in various CPT locations')

    fig.subplots_adjust(bottom=0.15, top=0.91)

    #ax.set_xticks(unique_ggm)

    ax.legend(fontsize=5, ncol=2)
    if filename:
        fig.savefig(filename, dpi=400)
    else:
        plt.show()

    count_df = DataFrame(ggm_counts.T, columns=columns)
    # Add a column to the left of the dataframe with the group names
    count_df.insert(0, 'CPT Location', rows)
    if filename:
        count_df.to_excel(filename + '.xlsx', index=False)
        # save count_df as a latex table with centering
        string = count_df.to_latex(escape=False, index=False)
        string = string.replace('\\toprule', '\\toprule\n\\centering')
        string = string.replace('\\midrule', '\\midrule\n\\centering')
        string = string.replace('\\bottomrule', '\\bottomrule\n\\centering')
        with open(filename + '.tex', 'w') as f:
            f.write(string)

    plt.close(fig)

def collection_of_plots(mgroup, fold):
    with open('string.txt', 'r') as f:
        string = f.read()
    string = string.replace('{mgroup}', mgroup).replace('{fold}', str(fold))
    return string

def loss_statistics(mgroups, distances):
    """This function takes the MSE loss for reconstruction and decoder of all models in mgroups,
    creates a correlation matrix and saves it as a figure. It also creates a scatterplot of the two losses"""

    from pandas import DataFrame, read_excel
    from matplotlib import pyplot as plt
    import matplotlib.patheffects as path_effects
    import seaborn as sns

    title_fontsize = 20
    xlabel_fontsize = 16
    ylabel_fontsize = 16

    destination_folder = 'Assignment figures/'
    ultimate_df = DataFrame(columns=['MAE_decoder', 'MAE_Rec', 'Distance'])
    ultimate_df_test = DataFrame(columns=['MAE_decoder', 'MAE_Rec', 'Distance'])

    for mgroup, distance in zip(mgroups, distances):
        df = read_excel(destination_folder + 'Renewed_figures/' + mgroup + '/loss_dict_new.xlsx')
        df_test = read_excel(destination_folder + 'Renewed_figures/' + mgroup + '/loss_dict_new_test.xlsx')

        dec_mse = df['MAE_decoder']
        rec_mse = df['MAE_Rec']
        dec_mse_test = df_test['MAE_decoder']
        rec_mse_test = df_test['MAE_Rec']
        dist = np.ones_like(dec_mse)*distance

        tmp_df = DataFrame({'MAE_decoder': dec_mse, 'MAE_Rec': rec_mse, 'Distance': dist})
        ultimate_df = ultimate_df.append(tmp_df, ignore_index=True)
        del tmp_df
        tmp_df = DataFrame({'MAE_decoder': dec_mse_test, 'MAE_Rec': rec_mse_test, 'Distance': dist})
        ultimate_df_test = ultimate_df_test.append(tmp_df, ignore_index=True)
    
    # Create a table with the correlation coefficient and spearman rank correlation coefficient for MAE_decoder and MAE_Rec for every distance
    corr_df = DataFrame(columns=['n', 'Correlation coefficient', 'Spearman rank correlation coefficient'])
    for distance in distances:
        tmp_df = ultimate_df[ultimate_df['Distance'] == distance]
        corr = tmp_df.corr(method='pearson')
        s_corr = tmp_df.corr(method='spearman')
        corr_df = corr_df.append(DataFrame({'n': len(tmp_df), 'Correlation coefficient': corr.iloc[0,1], 'Spearman rank correlation coefficient': s_corr.iloc[0,1]}, index=[distance]))
    
    corr_df.to_excel(destination_folder + 'Renewed_figures/Correlation_coefficients.xlsx')

    # Create correlation matrix
    corr = ultimate_df.corr()
    # mask = np.triu(np.ones_like(corr, dtype=bool))
    fig, ax = plt.subplots(figsize=(4,4))
    im = sns.heatmap(corr, cmap='RdYlBu', cbar=False)
    ticklabels = ['Decoder', 'Reconstruction', 'Distance']
    im.set_xticklabels(ticklabels, fontsize=8)
    im.set_yticklabels(ticklabels, fontsize=8)

    # Add the correlation coefficients to the heatmap
    for i in range(len(corr)):
        for j in range(len(corr)):
            text = ax.text(j+0.5, i+0.5, '{:.2f}'.format(corr.iloc[i,j]), ha='center', va='center', color='w', path_effects=[path_effects.Stroke(linewidth=1, foreground='black'), path_effects.Normal()], fontsize=10)


    ax.set_title('Correlation matrix of MAE losses and distance\ntraining set', fontsize=10)
    fig.subplots_adjust(right = 0.95, left=0.1, bottom=0.1)
    fig.savefig(destination_folder + 'Correlation_matrix.png', dpi=400)
    plt.close(fig)

    # Create a seaborn boxplot of the decoder loss to distance, with distance on the x-axis
    fig, ax = plt.subplots(figsize=(8, 6))

    models = mgroups
    for i, m in enumerate(models):
        if m == 'AVY':
            models[i] = 'AVY (+AWB)'

    im = sns.boxplot(x=ultimate_df['Distance'], y=ultimate_df['MAE_decoder'], ax=ax, orient='v')
    ax.set_xlabel('Max distance [m]', fontsize=xlabel_fontsize)
    ax.set_ylabel('MAE decoder', fontsize=ylabel_fontsize)
    ax.set_title('Boxplot of MAE decoder loss to distance\ntraining set', fontsize=title_fontsize)
    # Make the labels of the boxes model names
    for i, model in enumerate(models):
        im.patches[i].set_label(model)
    
    # Get the color of the patches
    colors = [patch.get_facecolor() for patch in im.patches]

    ax.legend(fontsize=12)

    fig.subplots_adjust(right = 0.95)

    fig.savefig(destination_folder + 'Boxplot_dec_dist.png', dpi=400)
    plt.close(fig)

    # Create scatterplot
    fig, ax = plt.subplots(figsize=(8, 6))
    c = np.ones((ultimate_df['Distance'].shape[0], 4))
    for i, distance in enumerate(distances):
        w = np.where(ultimate_df['Distance'] == distance)
        c[w] = colors[i]

    im = ax.scatter(ultimate_df['MAE_decoder'], ultimate_df['MAE_Rec'], c=c, zorder=3)
    ax.grid(True, 'both', zorder=1)
    ax.set_xlabel('Decoder', fontsize=xlabel_fontsize)
    ax.set_ylabel('Reconstruction', fontsize=ylabel_fontsize)
    ax.set_title('Scatter plot of MAE losses\ntraining set', fontsize=title_fontsize)
    # Add a legend grouping the models by color
    for i in range(len(models)):
        ax.scatter([], [], c=colors[i], label=models[i])
    ax.legend(fontsize=12)


    fig.subplots_adjust(right = 0.95)


    fig.savefig(destination_folder + 'Scatterplot.png', dpi=400)
    plt.close(fig)


    threshold = 0.6

    # Create a pie chart of the distance classes which have a MAE reconstruction loss above the threshold
    fig, ax = plt.subplots(figsize=(6, 5))
    # Get the number of distances above the threshold
    above = ultimate_df[ultimate_df['MAE_Rec'] > threshold]

    d = above['Distance'].value_counts()
    # colors = np.array(colors)
    # models = np.array(models)
    labels = []
    clrs  = []
    rnge = np.arange(len(distances))
    for dst in d.index:
        
        i = [i for i, d in enumerate(distances) if d == dst][0]
        clrs.append(colors[i])
        labels.append(models[i])
    # Create a pie chart
    im = ax.pie(d, labels=labels, startangle=90, counterclock=False, colors=clrs, wedgeprops={'linewidth': 1, 'edgecolor': 'black'})
    ax.set_title('Models with reconstruction MAE >0.6', fontsize=title_fontsize)
    fig.subplots_adjust(bottom=0.05)
    fig.savefig(destination_folder + 'Pie_chart.png', dpi=400)
    plt.close(fig)


    # Create the same figures for the test set
    corr = ultimate_df_test.corr()
    fig, ax = plt.subplots(figsize=(4,4))
    im = sns.heatmap(corr, cmap='RdYlBu', cbar=False)
    ticklabels = ['Decoder', 'Reconstruction', 'Distance']
    im.set_xticklabels(ticklabels, fontsize=8)
    im.set_yticklabels(ticklabels, fontsize=8)

    # Add the correlation coefficients to the heatmap
    for i in range(len(corr)):
        for j in range(len(corr)):
            text = ax.text(j+0.5, i+0.5, '{:.2f}'.format(corr.iloc[i,j]), ha='center', va='center', color='w', path_effects=[path_effects.Stroke(linewidth=1, foreground='black'), path_effects.Normal()], fontsize=10)
    ax.set_title('Correlation matrix of MAE losses and distance\ntest set', fontsize=10)
    fig.subplots_adjust(right = 0.95, left=0.1, bottom=0.1)
    fig.savefig(destination_folder + 'Correlation_matrix_test.png', dpi=400)
    plt.close(fig)

    # Create a seaborn boxplot of the decoder loss to distance, with distance on the x-axis
    fig, ax = plt.subplots(figsize=(8, 6))
    im = sns.boxplot(x=ultimate_df_test['Distance'], y=ultimate_df_test['MAE_decoder'], ax=ax, orient='v')
    ax.set_xlabel('Max distance [m]', fontsize=xlabel_fontsize)
    ax.set_ylabel('MAE decoder', fontsize=ylabel_fontsize)
    ax.set_title('Boxplot of MAE decoder loss to distance\ntest set', fontsize=title_fontsize)
    # Make the labels of the boxes model names
    for i, model in enumerate(models):
        im.patches[i].set_label(model)

    # Get the color of the patches
    colors = [patch.get_facecolor() for patch in im.patches]
    
    ax.legend(fontsize=12)

    fig.subplots_adjust(right = 0.95)

    fig.savefig(destination_folder + 'Boxplot_dec_dist_test.png', dpi=400)
    plt.close(fig)

    # Create scatterplot, with the values above the threshold highlighted
    fig, ax = plt.subplots(figsize=(8, 6))
    c = np.ones((ultimate_df_test['Distance'].shape[0], 4))
    for i, distance in enumerate(distances):
        w = np.where(ultimate_df_test['Distance'] == distance)
        c[w] = colors[i]
    
    test_below = ultimate_df_test[ultimate_df['MAE_Rec'] <= threshold]
    c_below = c[ultimate_df['MAE_Rec'] <= threshold]
    test_above = ultimate_df_test[ultimate_df['MAE_Rec'] > threshold]
    c_above = c[ultimate_df['MAE_Rec'] > threshold]

    im = ax.scatter(test_below['MAE_decoder'], test_below['MAE_Rec'], c=c_below, zorder=3)
    im = ax.scatter(test_above['MAE_decoder'], test_above['MAE_Rec'], c=c_above, marker='^', zorder=3)

    ax.grid(True, 'both', zorder=1)
    ax.set_xlabel('Decoder', fontsize=xlabel_fontsize)
    ax.set_ylabel('Reconstruction', fontsize=ylabel_fontsize)
    ax.set_title('Scatter plot of MAE losses\ntest set', fontsize=title_fontsize)
    # Add a legend grouping the models by color
    for i in range(len(models)):
        ax.scatter([], [], c=colors[i], label=models[i])
    ax.legend(fontsize=12)
    fig.subplots_adjust(right = 0.95)


    fig.savefig(destination_folder + 'Scatterplot_test.png', dpi=400)
    plt.close(fig)


    threshold = 0.6

    # Create a pie chart of the distance classes which have a MAE reconstruction loss above the threshold
    fig, ax = plt.subplots(figsize=(6, 5))
    # Get the number of distances above the threshold
    above = ultimate_df_test[ultimate_df_test['MAE_Rec'] > threshold]

    d = above['Distance'].value_counts()
    # colors = np.array(colors)
    # models = np.array(models)
    labels = []
    clrs  = []
    rnge = np.arange(len(distances))
    for dst in d.index:
            i = [i for i, d in enumerate(distances) if d == dst][0]
            clrs.append(colors[i])
            labels.append(models[i])
    # Create a pie chart
    im = ax.pie(d, labels=labels, startangle=90, counterclock=False, colors=clrs, wedgeprops={'linewidth': 1, 'edgecolor': 'black'})
    ax.set_title('Models with reconstruction MAE >0.6', fontsize=title_fontsize)
    fig.subplots_adjust(bottom=0.05)
    fig.savefig(destination_folder + 'Pie_chart_test.png', dpi=400)
    plt.close(fig)

def create_specific_plot(cpt_loc, depth_range, parameter):
    from Feat_aug import get_cpt_name
    import lasio
    cpt_name = get_cpt_name(cpt_loc)

    # load the data
    filename = f'../OneDrive - NGI/Documents/NTNU/MSC_DATA/combined/{cpt_name}.las'
    las = lasio.read(filename)
    depth = las['DEPT']
    parameter = las[parameter]

    # Find the indices of the depth range
    idxs = np.where((depth >= depth_range[0]) & (depth <= depth_range[1]))[0]

    # Create a plot of the parameter
    fig, ax = plt.subplots(figsize=(3, 5))
    ax.plot(parameter[idxs], depth[idxs], c='g')
    ax.set_xlabel('$q_c$ [MPa]')

    ax.invert_yaxis()

    ax.set_ylabel('Depth [meters below sea floor]')
    ax.set_title(f'{cpt_name} BH-CPT')
    fig.tight_layout()
    # fig.subplots_adjust(left=0.2, bottom=0.08, top=0.925)
    fig.savefig(f'Assignment figures/{cpt_name}_qc_stroke.png', dpi=350)
    plt.close(fig)

def count_char(files, not_used, typ='files'):
    """Counts the number of characters in a list of files"""
    count = 0
    for file in files:
        with open(file, 'r') as f:
            count += len(f.read())
    
    n_count = 0
    for file in not_used:
        with open(file, 'r') as f:
            n_count += len(f.read())
    
    # Create a pie chart with the number of characters in the files
    fig, ax = plt.subplots(figsize=(6, 5))
    im = ax.pie([count, n_count], labels=['Used', 'Not used'], startangle=90, counterclock=False, colors=[msc_color, coral], wedgeprops={'linewidth': 1, 'edgecolor': 'black'})
    ax.set_title('Number of characters in {}'.format(typ), fontsize=20)
    # Add the counts to the pie chart
    for i in range(len(im[0])):
        im[0][i].set_label('{:} characters'.format([count, n_count][i]))
    ax.legend(fontsize=12)
    fig.subplots_adjust(bottom=0.05)
    fig.savefig('Assignment figures/Number_of_characters_{}.png'.format(typ), dpi=400)
    plt.close(fig)
    return count, n_count
    
def count_characters_in_files():
    files = ['Scripts/init.py', 'Scripts/Feat_aug.py', 'Scripts/Log.py', 'Scripts/Architectures.py', 'Scripts/misc.py']
    not_used = ['Scripts/draft.py']
    code_count, code_ncount = count_char(files, not_used, typ='code')

    tempdir = r"C:\Users\SjB\Downloads\GEOL3095 Master's Assignment" + '/'
    files = ['Section 1/0. Title Page.tex', 'Section 1/1. Abstract.tex', 'Section 1/2. Sammendrag.tex', 'Section 1/3. Preface.tex', 'Section 1/7. Abbreviations.tex', 
             'Section 2/1. Introduction.tex', 'Section 2/2. Theory.tex', 'Section 2/3. Background.tex', 'Section 2/4. Method.tex', 'Section 2/5. Results.tex', 'Section 2/6. Discussion.tex', 'Section 2/7. Conclusions.tex',
             'Section 3/Appendix 1.tex']
    files = [tempdir+file for file in files]
    not_used = [tempdir+'Section 2/kladd.tex']
    thesis_count, thesis_ncount = count_char(files, not_used, typ='thesis')

    # Create a pie chart with the number of characters in the files
    fig, ax = plt.subplots(figsize=(6, 5))
    im = ax.pie([thesis_count, thesis_ncount, code_count, code_ncount], labels=['Thesis used', 'Thesis not used', 'Code used', 'Code not used'], startangle=90, counterclock=False, colors=[msc_color, dark_msc_color, coral, dark_coral], wedgeprops={'linewidth': 1, 'edgecolor': 'black'})
    ax.set_title('Number of characters in thesis and code', fontsize=20)
    # Add the counts to the pie chart
    for i in range(len(im[0])):
        im[0][i].set_label('{:} characters'.format([thesis_count, thesis_ncount, code_count, code_ncount][i]))
    ax.legend(fontsize=12)
    fig.subplots_adjust(bottom=0.05)
    fig.savefig('Assignment figures/Number_of_characters_thesis_and_code.png', dpi=400)
    plt.close(fig)




def renew_figs(mgroups, file_destination=r'./Assignment figures/Renewed_figures/', plot=True):
    """Based on the models in the argument, this function creates updated figures for the assignment"""
    from Feat_aug import create_full_trace_dataset, create_sequence_dataset, create_loo_trace_prediction_GGM, prediction_scatter_plot, plot_latent_space, predict_encoded_tree, get_cpt_name, get_cpt_data_scaler
    from sklearn.multioutput import MultiOutputRegressor
    from sklearn.ensemble import RandomForestRegressor
    from lightgbm import LGBMRegressor
    from pandas import concat

    for mgroup in mgroups:
        m_folder = 'Models/{}'.format(mgroup)
        destination_folder = file_destination + mgroup
        Path(destination_folder).mkdir(parents=True, exist_ok=True)

        with open(m_folder + '/param_dict.json') as f:
            param_dict = json.load(f)
        
        temp_param_dict = param_dict['dataset'].copy()
        temp_param_dict.pop('sequence_length')
        temp_param_dict.pop('stride')
        rf_params = param_dict['RF']
        lgbm_params = param_dict['LGBM']

        # Data
        Full_args = create_full_trace_dataset(**temp_param_dict)
        X, y, groups, nan_idxs, no_nan_idxs, _, _, GGM, _, minmax = Full_args

        args = create_sequence_dataset(**param_dict['dataset'])
        # X, y, groups, Z, GGM, GGM_unc
        X_seq = args[0]
        y_seq = args[1]
        GGM_seq = args[4]
        groups_seq = args[2]


        if plot: bar_plot_ggm(GGM_seq, groups_seq, filename=destination_folder+'/GGM_cptloc_distribution')

        scaler = get_cpt_data_scaler()
        scaled_back_y = np.ones_like(y_seq)
        for i in range(len(y_seq)):
            scaled_back_y[i] = scaler.inverse_transform(y_seq[i])
        describe_data(X_seq, scaled_back_y, groups_seq, GGM_seq, mdir=destination_folder+'/')
        # del scaled_back_y, X_seq, y_seq

        if plot: plot_dontfit(destination_folder+'/CPT_not_in_dataset')

        # Make excel table of the folds with loss values
        make_loss = not ((Path(m_folder + '/loss_dict.json').exists())or(Path(destination_folder + '/loss_dict.json').exists()))
        make_loss = True
        if not make_loss:
            with open(destination_folder + '/loss_dict_extra.json') as f:
                loss_dict = json.load(f)
                loss_dict_latex(loss_dict, destination_folder=destination_folder+'/')
        else:
            loss_dict = {}
            loss_dict_test = {}

        COMP_df = DataFrame([], columns = ['Depth', 'GGM', 'GGM_uncertainty', 'TRUE_qc', 'true_qc', 'CNN_qc', 'RF_qc', 'LGBM_qc', 'TRUE_fs', 'true_fs', 'CNN_fs', 'RF_fs', 'LGBM_fs', 'TRUE_u2', 'true_u2', 'CNN_u2', 'RF_u2', 'LGBM_u2'])

        # iterate over the folders which are named after the k-fold
        for fold in Path(m_folder).glob('Fold*'):
            print('###############' + fold.name + '###############')
            fold_number = fold.name[4:]
            fold_folder = file_destination + mgroup + '/' + str(fold.name)
            Path(fold_folder).mkdir(parents=True, exist_ok=True)

            with open(fold / 'LOO_group.txt') as f:
                loo_group = int(f.read()[11:].strip('[]'))
            
            with open(fold_folder + '/LOO_group.txt', 'w') as f:
                f.write(str(loo_group))
            
            # iterate over the files in the folder
            modelnames = [i for i in fold.glob('*.h5')]
            # Pop encoder and reconstruct models from the list
            encname = modelnames.pop([i for i, j in enumerate(modelnames) if 'encoder' in str(j)][0])
            recname = modelnames.pop([i for i, j in enumerate(modelnames) if 'reconstruct' in str(j)][0])
            modelname = modelnames[0]

            model = load_model(modelname)
            encoder = load_model(encname)
            reconstruct = load_model(recname)

            if make_loss:
                reconstruct.compile(loss='mae', metrics=['mae', 'mse', 'cosine_similarity'])
                X_test_seq = X_seq[np.isin(groups_seq.astype(int), loo_group)]
                y_test_seq = y_seq[np.isin(groups_seq.astype(int), loo_group)]
                X_train_seq = X_seq[~np.isin(groups_seq.astype(int), loo_group)]
                y_train_seq = y_seq[~np.isin(groups_seq.astype(int), loo_group)]
                metrics = reconstruct.evaluate(X_train_seq, [y_train_seq, X_train_seq], verbose=0)
                metrics_test = reconstruct.evaluate(X_test_seq, [y_test_seq, X_test_seq], verbose=0)
                for j, metric in enumerate(reconstruct.metrics_names):
                    loss_dict[f'Fold{fold_number}_{metric}'] = metrics[j]
                for j, metric in enumerate(reconstruct.metrics_names):
                    loss_dict_test[f'Fold{fold_number}_{metric}'] = metrics_test[j]

            
            # Train and test data
            test_idxs = np.where(groups.astype(int)==loo_group)[0]

            X_test = X[test_idxs]
            y_test = y[test_idxs]
            GGM_test = GGM[test_idxs]
            nan_idxs_test = nan_idxs[test_idxs]
            no_nan_idxs_test = no_nan_idxs[test_idxs]
            minmax_test = (minmax[0][test_idxs], minmax[1][test_idxs])

            train_idxs = np.where(groups.astype(int)!=loo_group)[0]
            X_train = X[train_idxs]
            y_train = y[train_idxs]
            nan_idxs_train = nan_idxs[train_idxs]
            no_nan_idxs_train = no_nan_idxs[train_idxs]

            

            # Create RF and LGBM models
            rf = MultiOutputRegressor(RandomForestRegressor(**rf_params))
            lgbm = MultiOutputRegressor(LGBMRegressor(**lgbm_params))

            encoding = encoder.predict(X_train)[no_nan_idxs_train].reshape(-1, 16)
            rf.fit(encoding, y_train[no_nan_idxs_train].reshape(-1, 3))
            lgbm.fit(encoding, y_train[no_nan_idxs_train].reshape(-1, 3))

            # Predictions for comparison
            depth = np.arange(temp_param_dict['zrange'][0], temp_param_dict['zrange'][1], 0.1)
            depths = []
            for i in range(X_test.shape[0]):
                depths.append(depth.copy())
            depths = np.array(depths)
            depths = depths[no_nan_idxs_test]
            ann_pred = model.predict(X_test)[no_nan_idxs_test].reshape(-1, 3)
            rf_pred = predict_encoded_tree(encoder, rf, X_test)[no_nan_idxs_test].reshape(-1, 3)
            lgbm_pred = predict_encoded_tree(encoder, lgbm, X_test)[no_nan_idxs_test].reshape(-1, 3)

            # Inverse transform the predictions and true values
            ann_pred = scaler.inverse_transform(ann_pred)
            rf_pred = scaler.inverse_transform(rf_pred)
            lgbm_pred = scaler.inverse_transform(lgbm_pred)


            # inverse transform the true values
            MMTEST = scaler.inverse_transform(minmax_test[1][no_nan_idxs_test].reshape(-1, 3))
            mmtest = scaler.inverse_transform(minmax_test[0][no_nan_idxs_test].reshape(-1, 3))

            TRUE_qc, TRUE_fs, TRUE_u2 = MMTEST[:, 0], MMTEST[:, 1], MMTEST[:, 2]
            true_qc, true_fs, true_u2 = mmtest[:, 0], mmtest[:, 1], mmtest[:, 2]

            ggm = GGM_test[no_nan_idxs_test]
            ggm_unc = np.zeros_like(ggm)

            

            
            comp = {'Depth': depths.flatten(), 'GGM': ggm.flatten(), 'GGM_uncertainty': ggm_unc.flatten(),
                    'TRUE_qc': TRUE_qc.flatten(), 'true_qc': true_qc.flatten(), 'CNN_qc': ann_pred[:, 0], 'RF_qc': rf_pred[:, 0], 'LGBM_qc': lgbm_pred[:, 0],
                    'TRUE_fs': TRUE_fs.flatten(), 'true_fs': true_fs.flatten(), 'CNN_fs': ann_pred[:, 1], 'RF_fs': rf_pred[:, 1], 'LGBM_fs': lgbm_pred[:, 1],
                    'TRUE_u2': TRUE_u2.flatten(), 'true_u2': true_u2.flatten(), 'CNN_u2': ann_pred[:, 2], 'RF_u2': rf_pred[:, 2], 'LGBM_u2': lgbm_pred[:, 2]}

            COMP_df = concat([COMP_df, DataFrame(comp)], ignore_index=True)

            # Make new figures
            latent_space_name = 'Latent_space_{}{}.png'.format(mgroup, fold_number)

            if plot:
                print('Creating latent space plot for {}_{}'.format(mgroup, fold_number))
                plot_latent_space(encoder,
                                latent_features=16,
                                X=X_test, 
                                valid_indices=no_nan_idxs_test, 
                                outside_indices=nan_idxs_test, 
                                GGM=GGM_test, 
                                filename= fold_folder + '/' + latent_space_name)

            if plot:
                for i in range(X_test.shape[0]):
                    extra = ''
                    if i > 0: extra = '_{}'.format(i)
                    for m, title in zip([model, [encoder, rf], [encoder, lgbm]], ['ANN', 'RF', 'LGBM']):
                        print('Creating prediction plots for {}_{}{} Fold {}'.format(title, mgroup, extra, fold_number))
                        create_loo_trace_prediction_GGM(m,
                                                        np.expand_dims(X_test[i], axis=0),
                                                        np.expand_dims(y_test[i], axis=0),
                                                        np.expand_dims(GGM_test[i], axis=0),
                                                        zrange=temp_param_dict['zrange'],
                                                        filename=fold_folder + '/pred_' + title + '_{}{}{}.png'.format(mgroup, fold_number, extra),
                                                        title=title + ' ' + get_cpt_name(loo_group) + extra.replace('_', ' '),
                                                        minmax=(np.expand_dims(minmax_test[0][i], axis=0), np.expand_dims(minmax_test[1][i], axis=0)))
                        
                        print('Creating scatter plots for {}_{}{} Fold {}'.format(title, mgroup, extra, fold_number))
                        prediction_scatter_plot(m,
                                                np.expand_dims(X_test[i], axis=0),
                                                np.expand_dims(y_test[i], axis=0),
                                                filename=fold_folder + '/scatter_' + title + '_{}{}{}.png'.format(mgroup, fold_number, extra),
                                                title=title + ' ' + get_cpt_name(loo_group) + extra.replace('_', ' '),
                                                bins=15)


                # Make illustration of the reconstruction model
                print('Creating reconstruction plot for {} Fold {}'.format(mgroup, fold_number))
                plot_reconstruction(reconstruct, filename=fold_folder + '/loo_reconstruction_{}{}.png'.format(mgroup, fold_number), cpt_loc=loo_group)
        # Dump the COMP_df to a pickle file
        COMP_df.to_pickle(file_destination + mgroup + '/COMP_df.pkl')

        if make_loss:
            with open(destination_folder + '/loss_dict_extra.json', 'w') as f:
                json.dump(loss_dict, f, indent=4)
            loss_dict_latex(loss_dict, destination_folder + '/', filename='loss_dict_new')
            with open(destination_folder + '/loss_dict_extra_test.json', 'w') as f:
                json.dump(loss_dict_test, f, indent=4)
            loss_dict_latex(loss_dict_test, destination_folder + '/', filename='loss_dict_new_test')

        # Make best and worst case cross validation excel file
        make_cv_excel_bestworst(filename=file_destination + mgroup + '/CV.xlsx', COMP_DF=COMP_df)


if __name__ == '__main__':
    # import pyperclip
    # nums = [8]
    # for n in nums:
    #     print(n)
    #     pyperclip.copy(collection_of_plots('AVX', n))
    #     input('Press enter to continue')
    
    # create_specific_plot(cpt_loc=45, depth_range=[40, 60], parameter='QC')

    count_characters_in_files()

    # renew_figs(['AVW', 'AVX', 'AVY', 'AWD'], plot=False)
    # loss_statistics(['AVW', 'AVX', 'AVY', 'AWD'], [0.4, 1, 5, 10])

    # m = ['AVW', 'AVX', 'AVY', 'AWD']
    # d = 'Assignment Figures/Renewed_figures/'


    # for model in m:
    #     with open(d + model + '/loss_dict_extra.json', 'r') as f:
    #         loss_dict = json.load(f)
    #     loss_dict_latex(loss_dict, d + model + '/')
    
    # loss_statistics(['AVW', 'AVX', 'AVY', 'AWD'], [0.4, 1, 5, 10])
    # import pickle
    # with open(r'./Assignment figures/Renewed_figures/AVY/COMP_df.pkl', 'rb') as f:
    #     COMP_df = pickle.load(f)



    # make_cv_excel_bestworst(filename=r'./Assignment figures/Renewed_figures/AVY/CV_x.xlsx', COMP_DF=COMP_df)

    # from Feat_aug import create_full_trace_dataset
    # full_args = create_full_trace_dataset()
    # ggm = full_args[7]

    # cmap, norm, _ = get_GGM_cmap(ggm)

    # umap = get_umap_func()
    # # Create a dataframe with the hex values of the GGM
    # cdict = {}
    # for g in np.unique(ggm):
    #     cdict[g] = cmap(norm(g))[:3]

    # Print out the hex values for matplotlib colors 'g', 'orange' and 'b'
    # import matplotlib.colors as mcolors
    # print(mcolors.to_hex('g'))
    # print(mcolors.to_hex('orange'))
    # print(mcolors.to_hex('b'))

    s = ''
    for i in range(16):
        s += f'c{i+1} '
    print(s)


    
